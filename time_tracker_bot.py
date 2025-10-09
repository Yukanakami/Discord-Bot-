import discord 
from discord.ext import commands
import asyncio
import datetime
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


"""
Discord Time Tracking Bot
Copyright (c) 2023 Your Name. All rights reserved.

This bot allows users to track their work hours and breaks through Discord commands.
Features include clocking in/out, break tracking, status checks, and Excel report generation.
"""

intents = discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix='!', intents=intents)

    
DATA_FILE = 'time_data.json'

ALLOWED_RESET_USERS = ['USER_ID_1', 'USER_ID_2']  # Replace with actual user IDs

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return {}
    
def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=4)

time_data = load_data()

@bot.event 
async def on_ready():
    print(f'{bot.user} has connected to Discord!')

@bot.command(name='clock')
async def clock_command(ctx, action=None):
    user_id = str(ctx.author.id)
    user_name = str(ctx.author)

    if user_id not in time_data:
        time_data[user_id] = {
            'name': user_name,
            'clock_in': None,
            'clock_out': None,
            'breaks': []
        }

    if action is None:
        
        await ctx.send("Please specify 'in' or 'out'.", ephemeral=True)
        return

    if action.lower() == 'in':
        if time_data[user_id]['clock_in'] is not None:
            await ctx.send("You are already clocked in today", ephemeral=True)
            return
        
        time_data[user_id]['clock_in'] = datetime.datetime.now().isoformat()
        save_data(time_data)
       
        await ctx.send(f"✅ {ctx.author.mention} clocked in at {datetime.datetime.now().strftime('%H:%M:%S')}")

    elif action.lower() == 'out':
        if time_data[user_id]['clock_in'] is None:
            await ctx.send("You need to clock in first!", ephemeral=True)
            return

        if time_data[user_id]['clock_out'] is not None:
            await ctx.send("You have already clocked out today.", ephemeral=True)
            return

        current_breaks = [b for b in time_data[user_id]['breaks'] if b.get('end') is None]
        if current_breaks:
            await ctx.send("You're currently on a break. Please end your break before clocking out.", ephemeral=True)
            return
        
        time_data[user_id]['clock_out'] = datetime.datetime.now().isoformat()
        save_data(time_data)
        
        await ctx.send(f"✅ {ctx.author.mention} clocked out at {datetime.datetime.now().strftime('%H:%M:%S')}")

    else:
        await ctx.send("Invalid action. Please use `!clock in` or `!clock out`", ephemeral=True)

@bot.command(name='break')
async def break_command(ctx, action=None):
    user_id = str(ctx.author.id)
    user_name = str(ctx.author)

    if user_id not in time_data or time_data[user_id]['clock_in'] is None:
        await ctx.send("You need to clock in first!", ephemeral=True)
        return

    if action is None:
        await ctx.send("Please specify an action: `!break start` or `!break end`", ephemeral=True)
        return

    if action.lower() == 'start':
        current_breaks = [b for b in time_data[user_id]['breaks'] if b.get('end') is None]     
        if current_breaks:
            await ctx.send("You've already started a break!", ephemeral=True)
            return
        
        if time_data[user_id]['clock_out'] is not None:
            await ctx.send("You have already clocked out today.", ephemeral=True)
            return
        
        time_data[user_id]['breaks'].append({
            'start': datetime.datetime.now().isoformat(),
            'end': None
        })
        save_data(time_data)
     
        await ctx.send(f"☕ {ctx.author.mention} started a break at {datetime.datetime.now().strftime('%H:%M:%S')}")

    elif action.lower() == 'end':
        current_breaks = [b for b in time_data[user_id]['breaks'] if b.get('end') is None]
        if not current_breaks:
            await ctx.send("You're not currently on a break!", ephemeral=True)
            return

        current_breaks[0]['end'] = datetime.datetime.now().isoformat()
        save_data(time_data)
        
        await ctx.send(f"✅ {ctx.author.mention} ended the break at {datetime.datetime.now().strftime('%H:%M:%S')}")

    else:
        await ctx.send("Invalid action. Use `!break start` or `!break end`", ephemeral=True)

@bot.command(name='status')
async def status_command(ctx):
    user_id = str(ctx.author.id)

    if user_id not in time_data or time_data[user_id]['clock_in'] is None:
        await ctx.send("You haven't clocked in yet!", ephemeral=True)
        return
    
    status_message = f"**Status for {ctx.author.mention}:**\n"
    status_message += f"⏰ Clocked in: {datetime.datetime.fromisoformat(time_data[user_id]['clock_in']).strftime('%H:%M:%S')}\n" 

    if time_data[user_id]['clock_out']:
        status_message += f"🚪 Clocked out: {datetime.datetime.fromisoformat(time_data[user_id]['clock_out']).strftime('%H:%M:%S')}\n"

    current_breaks = [b for b in time_data[user_id]['breaks'] if b.get('end') is None] 
    if current_breaks:
        status_message += f"☕ On a break since {datetime.datetime.fromisoformat(current_breaks[0]['start']).strftime('%H:%M:%S')}\n"
    else:
        status_message += "✅ Not currently on a break\n"

    await ctx.send(status_message, ephemeral=True)

@bot.command(name='report')
async def report_command(ctx):
    if not any(data['clock_in'] is not None for data in time_data.values()):
        await ctx.send("No time data available to generate a report.", ephemeral=True)
        return
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Tracking Report"

    headers = ["User", "Clock In", "Clock Out", "Break Start", "Break End", "Total Break Time", "Total Work Time"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row = 2
    for user_id, data in time_data.items():
        if data['clock_in'] is None:
            continue

        clock_in = datetime.datetime.fromisoformat(data['clock_in'])
        clock_out = datetime.datetime.fromisoformat(data['clock_out']) if data['clock_out'] else None

        total_break_time = datetime.timedelta(0)
        for break_entry in data['breaks']:
            if break_entry['end']:
                start = datetime.datetime.fromisoformat(break_entry['start'])
                end = datetime.datetime.fromisoformat(break_entry['end'])
                total_break_time += (end - start)

        if clock_out:
            total_work_time = (clock_out - clock_in) - total_break_time
        else:
            total_work_time = "Still working"

        break_starts = [datetime.datetime.fromisoformat(b['start']).strftime('%H:%M') for b in data['breaks']]
        break_ends = [datetime.datetime.fromisoformat(b['end']).strftime('%H:%M') if b['end'] else "Ongoing" for b in data['breaks']]

        ws.cell(row=row, column=1, value=data['name'])
        ws.cell(row=row, column=2, value=clock_in.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=3, value=clock_out.strftime('%Y-%m-%d %H:%M') if clock_out else "Not clocked out")
        ws.cell(row=row, column=4, value=", ".join(break_starts) if break_starts else "No breaks")
        ws.cell(row=row, column=5, value=", ".join(break_ends) if break_ends else "No breaks")
        ws.cell(row=row, column=6, value=str(total_break_time))
        ws.cell(row=row, column=7, value=str(total_work_time) if isinstance(total_work_time, datetime.timedelta) else total_work_time)

        row += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    filename = f"time_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    await ctx.send("Here's the time tracking report:", file=discord.File(filename))

    os.remove(filename)

@bot.command(name="reset")
async def reset_command(ctx):
    if str(ctx.author.id) not in ALLOWED_RESET_USERS:
        await ctx.send("You don't have permission to use this command.", ephemeral=True)
        return
    
    def check(m):
        return m.author == ctx.author and m.content.lower() == 'confirm'
    
    await ctx.send("Are you sure you want to reset all time data? This cannot be undone. Type 'confirm' to proceed.", ephemeral=True)

    try:
        await bot.wait_for('message', timeout=30.0, check=check)
    except asyncio.TimeoutError:
        await ctx.send("Reset cancelled.", ephemeral=True)
        return
    
    global time_data
    time_data = {}
    save_data(time_data)
    await ctx.send("All time data has been reset.")

if name == "main":
    
    bot.run('Your_bot_token_here')
